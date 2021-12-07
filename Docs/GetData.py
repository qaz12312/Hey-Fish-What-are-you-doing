#! /usr/bin/python3.7
"""
write data into .csv
"""
import argparse
import os
import numpy
import openpyxl
import pandas
import re

DIR_PATH = os.path.split(os.path.abspath(__file__))[0] + '/'

def isset(v): 
    try : 
        type (eval(v)) 
    except : 
        return  0  
    else : 
        return  1

def result_txt_to_excel(filename, execl_name):
    with open(filename, 'r') as f:
        content_list = f.readlines()
    
    # 建立 Excel 活頁簿
    wb = openpyxl.Workbook()
    # 建立工作表 & 設定工作表名稱
    sheet_dict = dict()
    sheet_dict['acc'] = wb.create_sheet("準確度", 0)
    sheet_dict['loss'] = wb.create_sheet("誤差值", 0)
    sheet_dict['time'] = wb.create_sheet("訓練時間", 0)

    # 將陣列寫入 Excel 工作表
    for sheet in sheet_dict.keys():
        sheet_dict[sheet].append(list(range(41)))

    row_i = 1
    for line in content_list:
        # strip 移出字串頭尾的換行
        line = line.strip('\n')
        if line[0] == '[':
            line = re.search(r'\[.*\]', line).group(0)
            value = "(" + ",".join(re.findall(r'\d+', line)) + ")"
            # 欄位
            col_j = 1
            # 行數遞增
            row_i += 1
            # 將值寫入 (row_i, col_j)
            for sheet in sheet_dict.keys():
                # 修改儲存格內容
                sheet_dict[sheet].cell(row=row_i, column=col_j, value=value)
        else:
            # 用''替換掉'\t'
            line = line.replace("\t","")
            # line = line.split(',')
            value_list = list()
            for regex in [r'(?<=Acc=)\d+\.?\d*', r'(?<=Loss=)\d+\.?\d*', r'(?<=time=)\d+\.?\d*']:
                value_list.append(re.search(regex, line).group(0))
            col_j = int(re.search(r'\[\d+ ?\]', line).group(0)[1:-1]) + 1
            # 將值寫入 (row_i, col_j)
            for sheet,value in zip(sheet_dict.keys(),value_list):
                # 修改儲存格內容
                sheet_dict[sheet].cell(row=row_i, column=col_j, value=value)

    # 儲存 Excel 活頁簿至檔案
    excel_path = DIR_PATH + execl_name + ".xlsx"
    wb.save(filename= excel_path)
    print("excel location: " + excel_path)


def row_to_excel(data_list, execl_name):
    '''
    Parameters
    ----------
    data_list : `list`
    execl_name : `str`

    Example
    ----------
    >>> row_to_excel([[1, 2, 3], [4, 5, 6]])
    | 1 | 2 | 3 |
    | 4 | 5 | 6 |
    '''
    # 建立 NumPy 陣列
    myArr = numpy.array(data_list)

    # 建立 Excel 活頁簿
    wb = openpyxl.Workbook()

    # 取得作用中的工作表
    sheet = wb.active

    # 設定工作表名稱
    sheet.title = execl_name

    # 將 NumPy 陣列寫入 Excel 工作表
    for x in myArr:
        sheet.append(x.tolist())

    # 儲存 Excel 活頁簿至檔案
    excel_path = DIR_PATH + execl_name + "_row1.xlsx"
    wb.save(filename= excel_path)
    print("excel location: " + excel_path)

def col_to_execl(data_dict,execl_name):
    '''
    Parameters
    ----------
    data_dict : `dict`
    execl_name : `str`

    Example
    ----------
    >>> col_to_execl({'col1': [1, 2, 3], 'col2': [4, 5, 6]})
    |  | col1 | col2 |
    |  |   1  |   4  |
    |  |   2  |   5  |
    |  |   3  |   6  |
    '''
    # 建立 Pandas 資料表
    df = pandas.DataFrame(data=data_dict)

    # 將 Pandas 資料表寫入 Excel 檔案
    excel_path = DIR_PATH + execl_name + "_col1.xlsx"
    df.to_excel(excel_path)
    print("excel location: " + excel_path)


parser = argparse.ArgumentParser(description="Store into excel.")
parser.add_argument('type', help='change_way',type=int)
args = parser.parse_args()
if __name__ == '__main__':
    #  data
    data = [
        list(range(1,41)),
        [0.512,	0.465,	0.512,	0.506,	0.584,	0.619,	0.488,	0.544,	0.600,	0.514,	0.468,	0.609,	0.626,	0.629,	0.569,	0.559,	0.564,	0.549,	0.619,	0.594,	0.612,	0.598,	0.529,	0.609,	0.610,	0.628,	0.649,	0.669,	0.617,	0.587,	0.614,	0.684,	0.640,	0.636,	0.650,	0.635,	0.579,	0.628,	0.669,	0.654,],
        [0.553,	0.540,	0.517,	0.458,	0.553,	0.549,	0.511,	0.525,	0.513,	0.566,	0.620,	0.651,	0.553,	0.644,	0.664,	0.612,	0.644,	0.621,	0.631,	0.647,	0.717,	0.658,	0.751,	0.675,	0.669,	0.714,	0.610,	0.715,	0.669,	0.647,	0.675,	0.667,	0.675,	0.701,	0.642,	0.713,	0.642,	0.627,	0.702,	0.707,],
        [0.637,	0.533,	0.406,	0.525,	0.622,	0.639,	0.527,	0.415,	0.617,	0.525,	0.680,	0.518,	0.602,	0.569,	0.652,	0.650,	0.644,	0.568,	0.669,	0.634,	0.592,	0.597,	0.576,	0.561,	0.598,	0.585,	0.565,	0.632,	0.622,	0.577,	0.614,	0.664,	0.777,	0.642,	0.703,	0.704,	0.568,	0.652,	0.627,	0.715,],
        [0.593,	0.614,	0.531,	0.576,	0.560,	0.586,	0.635,	0.753,	0.618,	0.580,	0.559,	0.620,	0.582,	0.641,	0.656,	0.585,	0.660,	0.644,	0.571,	0.589,	0.656,	0.647,	0.663,	0.594,	0.583,	0.633,	0.786,	0.667,	0.631,	0.718,	0.606,	0.648,	0.676,	0.707,	0.595,	0.753,	0.651,	0.757,	0.663,	0.741,],
        [0.482,	0.646,	0.641,	0.592,	0.564,	0.512,	0.633,	0.547,	0.616,	0.625,	0.608,	0.579,	0.606,	0.605,	0.617,	0.579,	0.557,	0.622,	0.647,	0.656,	0.621,	0.583,	0.656,	0.647,	0.630,	0.636,	0.614,	0.657,	0.617,	0.683,	0.615,	0.684,	0.690,	0.637,	0.636,	0.658,	0.617,	0.636,	0.679,	0.671,],
        [0.606,	0.721,	0.575,	0.485,	0.616,	0.652,	0.642,	0.732,	0.571,	0.690,	0.592,	0.678,	0.750,	0.656,	0.613,	0.639,	0.652,	0.649,	0.730,	0.688,	0.679,	0.673,	0.698,	0.680,	0.693,	0.647,	0.726,	0.643,	0.711,	0.676,	0.707,	0.657,	0.702,	0.724,	0.730,	0.668,	0.678,	0.718,	0.701,	0.719,],
        [0.528,	0.525,	0.567,	0.641,	0.567,	0.665,	0.524,	0.545,	0.563,	0.634,	0.611,	0.604,	0.700,	0.622,	0.607,	0.604,	0.618,	0.585,	0.603,	0.635,	0.668,	0.657,	0.685,	0.623,	0.732,	0.623,	0.713,	0.744,	0.604,	0.630,	0.659,	0.655,	0.602,	0.745,	0.656,	0.668,	0.631,	0.583,	0.652,	0.657,],
        [0.564,	0.652,	0.695,	0.473,	0.568,	0.704,	0.569,	0.707,	0.648,	0.687,	0.625,	0.748,	0.645,	0.586,	0.807,	0.647,	0.671,	0.667,	0.621,	0.672,	0.602,	0.670,	0.661,	0.702,	0.631,	0.660,	0.672,	0.761,	0.792,	0.754,	0.781,	0.597,	0.645,	0.652,	0.620,	0.737,	0.644,	0.687,	0.642,	0.735,],
        
       ]
    accs=[0.571 ,0.804 ,0.634 ,0.503 ,0.593 ,0.451 ,0.415 ,0.525 ,0.648 ,0.695 ,0.527 ,0.678 ,0.585 ,0.596 ,0.449 ,0.573 ,0.629 ,0.615 ,0.527 ,0.496 ,0.437 ,0.626 ,0.602 ,0.524 ,0.782 ,0.570 ,0.690 ,0.832 ,0.719 ,0.642 ,0.467 ,0.685 ,0.686 ,0.594 ,0.533]
    losses = [1.243 ,0.931 ,1.763 ,2.428 ,2.234 ,2.992 ,3.377 ,2.672 ,2.126 ,1.750 ,2.869 ,1.953 ,2.310 ,2.342 ,2.423 ,2.394 ,2.430 ,2.403 ,2.610 ,3.186 ,3.680 ,2.543 ,2.216 ,2.980 ,0.659 ,2.653 ,1.980 ,1.089 ,1.991 ,2.409 ,3.711 ,1.992 ,2.189 ,2.745 ,2.630]
    run_times=[ 15.650346279144287 ,15.78990650177002 ,16.00351595878601 ,15.892401456832886 ,16.952032804489136 ,17.230820655822754 ,17.90560007095337 ,17.961915731430054 ,19.014283418655396 ,19.52064347267151 ,20.55557608604431 ,20.983111143112183 ,22.42049741744995 ,22.967249631881714 ,23.50215435028076 ,23.83486843109131 ,25.25700283050537 ,25.84507155418396 ,26.65250873565674 ,27.14936137199402 ,28.21666431427002 ,29.121891975402832 ,28.762715578079224 ,29.29419231414795 ,31.608169078826904 ,32.489532232284546 ,33.43353581428528 ,34.550477743148804 ,35.59119415283203 ,36.82110071182251 ,37.26888656616211 ,37.65978455543518 ,39.520251750946045 ,40.6717472076416 ,42.080076932907104 ]
    if args.type == 1:
        row_to_excel(data,"hidden")
    elif args.type == 2:
        col_to_execl({'accs':accs,'losses':losses,'run_times':run_times},"hidden")
    elif args.type == 3:
        result_txt_to_excel('log/records/1000_all_result.txt', "All_hidden")
